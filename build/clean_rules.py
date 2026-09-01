#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
주문정합성체크 정제 파이프라인 (플랜 Task 1 + Task 2)

원본 xlsx 2종(주문에러체크 C10B2220 / 항목간주문에러체크 C10B2221)을 파싱·클렌징(C-1~C-7)하고
EAV 승격 후 병합 정책(M1~M9)으로 통합 룰셋 시드 JSON을 생성한다.

사용:
  python3 build/clean_rules.py --check            # 어서션 전부 실행 (실패 시 exit 1)
  python3 build/clean_rules.py --check --emit     # + build/order_consistency_seed.json 생성
  python3 build/clean_rules.py --inject           # modules/order-consistency.html 마커에 시드 주입(멱등)
  python3 build/clean_rules.py --order P --cross P  # 입력 경로 재지정

입력(필수): 원본 xlsx 2종 — 주문에러체크.xlsx / 항목간주문에러체크.xlsx.
저장소에 포함되지 않으므로 다음 중 하나로 공급한다.
  sources/ 디렉토리에 배치 (기본 탐색 경로) / OC_XLSX_DIR=<디렉토리> / --order·--cross

근거 문서: docs/research/expert1-data-quality.md §3 (C-1~C-13),
          docs/research/expert2-rule-semantics.md §2(M1~M9)·§3(스키마)·§4(평가 의미론),
          docs/specs/2026-09-01-order-consistency-design.md
편차 기록: docs/research/pipeline_notes.md
"""

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent

# 원본 xlsx 2종은 저장소에 포함되지 않는다(사내 기준정보 원본). 경로는 머신 독립적으로 푼다:
#   1) --order / --cross 로 명시  2) 환경변수 OC_XLSX_DIR  3) 저장소 sources/
# 어느 것도 없으면 sources/ 경로를 그대로 돌려줘 에러 메시지가 "어디에 두면 되는지"를 가리키게 한다.
# (절대경로 하드코딩은 신규 클론·다른 머신·CI에서 빌드 절차를 재현 불가로 만든다.)
SOURCE_DIR = ROOT / "sources"
ORDER_XLSX_NAME = "주문에러체크.xlsx"
CROSS_XLSX_NAME = "항목간주문에러체크.xlsx"


def default_source(name):
    import os

    env = os.environ.get("OC_XLSX_DIR")
    candidates = []
    if env:
        candidates.append(Path(env).expanduser() / name)
    candidates.append(SOURCE_DIR / name)
    for p in candidates:
        if p.exists():
            return str(p)
    return str(SOURCE_DIR / name)


DEFAULT_ORDER = default_source(ORDER_XLSX_NAME)
DEFAULT_CROSS = default_source(CROSS_XLSX_NAME)
SEED_OUT = ROOT / "build" / "order_consistency_seed.json"
MODULE_HTML = ROOT / "modules" / "order-consistency.html"
MARKER_START = "/*__OC_SEED_START__*/"
MARKER_END = "/*__OC_SEED_END__*/"

RAW_NUM_OPS = {">", ">=", "<", "<=", "BETWEEN1", "BETWEEN2"}
OP_MAP = {"=": "EQ", "!=": "NE", ">": "GT", ">=": "GE", "<": "LT", "<=": "LE"}
NULL_OPS = {"IS_NULL", "NOT_NULL"}

# ---------------------------------------------------------------------------
# 전용 컬럼(한국어 헤더) -> 영문 필드 키 (플랜 Task 1 Step 3 매핑)
# 편차 1건: 포장메세지코드는 플랜이 ORD_TEM_CD로 지정했으나 원본 데이터에서
# ORD_TEM_CD는 EAV '영업팀코드'(order#45/cross#33, A18 "영업팀코드 누락")로 실사용 중이라
# 충돌한다. 근거 없는 병합을 피하기 위해 임시 키 PAK_MSG_CD를 부여했다
# (EAV 근거 없음 — docs/research/pipeline_notes.md 편차 D-1).
COLUMN_KEY_MAP = {
    "품명코드": "PRD_NM_CD",
    "제품형태": "PRD_SHP",
    "주문환산두께": "ORD_EXC_THK",
    "주문환산폭": "ORD_EXC_WTH",
    "주문Slit조수": "ORD_SLIT_GRP_CNT",
    "최종수요가코드": "FNL_CUS_CD",
    "주문폭관리코드": "ORD_WTH_MNG_CD",
    "주문표면처리코드": "ORD_SUR_HND_CD",
    "주문내경링종류구분": "ORD_SLV_KND_TP",
    "주문용도코드": "ORD_USG_CD",
    "주문보호필름상세코드": "ORD_PTT_FLM_CD",
    "주문두께관리코드": "ORD_THK_MNG_CD",
    "포장메세지코드": "PAK_MSG_CD",  # 편차 D-1 (플랜: ORD_TEM_CD)
    "주문Edge지정구분": "ORD_EDG_ASG_TP",
    "EMBOSS무늬": "EMBS_CD",
    "주문포장단중하한값": "ORD_PAK_UNT_WGT_LLV",
    "주문포장단중상한값": "ORD_PAK_UNT_WGT_ULV",
    "도금량지정코드": "GW_ASG_CD",
    "주문Spangle구분": "ORD_SPNL_TP",
    "주문포장방법": "ORD_PAK_MTH",
}
KEY_COLUMN_LABEL = {v: k for k, v in COLUMN_KEY_MAP.items()}

# 라벨 미상 허용 키 (메시지에서 라벨 역추론 불가 — 코드 그대로 노출, 날조 금지)
LABEL_UNKNOWN_OK = {"RSN_TP_FRN"}

NUMBER_KEYWORDS = ("두께", "폭", "단중", "조수", "매수", "길이", "내경", "외경", "직경", "중량", "평량")
NUMBER_EXCLUDE_SUFFIX = ("코드", "구분", "방법", "단위", "약호", "번호")
CODE_LABEL_SUFFIX = ("코드", "구분", "방법", "무늬")
CODE_KEY_SUFFIX = ("_CD", "_TP", "_MTH", "_KND")
FIELD_UNITS = {
    "ORD_EXC_THK": "mm", "ORD_EXC_WTH": "mm",
    "ORD_PAK_UNT_WGT_LLV": "kg", "ORD_PAK_UNT_WGT_ULV": "kg",
    **{f"ORD_MIX_WTH{i}": "mm" for i in range(1, 11)},
}

# ---------------------------------------------------------------------------
# 값 코드 라벨 사전 — 전부 에러 메시지에서 역추론한 것만, certainty '추정'
# (지시: 근거 없는 라벨 날조 금지. '확정'은 화면에서 사용자 승격 시에만 부여)
VALUE_LABELS = (
    [{"field": "PRD_NM_CD", "code": "G", "label": "GI",
      "certainty": "추정", "evidence": "order#75/cross#63 'GI Spangle 선택 오류'(품명 G)"},
     {"field": "PRD_NM_CD", "code": "L", "label": "G/L",
      "certainty": "추정", "evidence": "order#76 'G/L Spangle 선택 오류'(품명 L)·order#26 'G/L 조관수지용'"},
     {"field": "PRD_NM_CD", "code": "3", "label": "CCGI",
      "certainty": "추정", "evidence": "cross#65 'CCGI Spangle 선택 오류'(품명 3 단독)"},
     {"field": "PRD_NM_CD", "code": "6", "label": "CCGI",
      "certainty": "추정", "evidence": "order#77 'CCGI Spangle 선택 오류'(품명 3,6)"},
     {"field": "PRD_NM_CD", "code": "4", "label": "CCLI",
      "certainty": "추정", "evidence": "cross#66 'CCLI Spangle 선택 오류'(품명 4 단독)"},
     {"field": "PRD_NM_CD", "code": "5", "label": "CCAI",
      "certainty": "추정", "evidence": "order#127 'CCAI 코일형'(품명 =5)·order#132 'CCAI'(품명 =5)"},
     {"field": "PRD_NM_CD", "code": "D", "label": "F/H",
      "certainty": "추정", "evidence": "order#48 'F/H 제품만 STEEL RING 가능'(품명 D 외 금지)"},
     {"field": "PRD_SHP", "code": "C", "label": "코일형",
      "certainty": "추정", "evidence": "order#95 '도금 코일형'(제품형태 =C) 등 다수"},
     {"field": "PRD_SHP", "code": "S", "label": "시트(SHEET)형",
      "certainty": "추정", "evidence": "order#29 'ZERO S/T SHEET 불가'(제품형태 =S) 등 다수"},
     {"field": "ORD_SLV_KND_TP", "code": "S", "label": "STEEL RING",
      "certainty": "추정", "evidence": "order#48 'F/H 제품만 STEEL RING 가능'(내경링 =S)"},
     {"field": "ORD_SLV_KND_TP", "code": "N", "label": "지관 미삽입",
      "certainty": "추정", "evidence": "order#85/#87 '지관삽입필수'(내경링 =N일 때 발화)"},
     {"field": "ORD_EDG_ASG_TP", "code": "M", "label": "Mill Edge",
      "certainty": "추정", "evidence": "order#34 'Mill Edge만 주문 가능'(C,N,S 금지 → 잔여 M)"},
     {"field": "ORD_EDG_ASG_TP", "code": "N", "label": "Zero S/T",
      "certainty": "추정", "evidence": "order#29 'ZERO S/T SHEET 불가'(Edge =N)"},
     {"field": "ORD_USG_CD", "code": "C15000", "label": "PE-FOAM 부착용",
      "certainty": "추정", "evidence": "order#8/#9 'PE-FOAM 부착용(C15000)'"},
     {"field": "ORD_USG_CD", "code": "G0691P", "label": "세탁기 앞판(COVER)",
      "certainty": "추정", "evidence": "order#11 '세탁기 앞판(COVER)'(용도 =G0691P)"},
     {"field": "ORD_USG_CD", "code": "I0101P", "label": "노트북 ODD용",
      "certainty": "추정", "evidence": "order#25 '노트북 ODD용 전용'(용도 I0101P 외 불가)"},
     {"field": "ORD_SUR_HND_CD", "code": "LP", "label": "G/L 조관수지용",
      "certainty": "추정", "evidence": "order#26 'G/L 조관수지용 주문표면처리코드'(=LP)"},
     {"field": "ORD_SUR_HND_CD", "code": "OD", "label": "노트북 ODD용 Non-Cr 수지",
      "certainty": "추정", "evidence": "order#25 '노트북 ODD용 전용 개발된 Non-Cr 수지'(=OD)"},
     {"field": "ORD_SUR_HND_CD", "code": "CX", "label": "이상재처리용",
      "certainty": "추정", "evidence": "order#28 '이상재처리용 - 표면처리코드 오류'(=CX)"},
     {"field": "ORD_SUR_HND_CD", "code": "X2", "label": "외장재용",
      "certainty": "추정", "evidence": "order#17 '후처리 X2(외장재용)'"},
     {"field": "ORD_PAK_MTH", "code": "AXXX00", "label": "포장 없음",
      "certainty": "추정", "evidence": "order#40/cross#28 '포장 없음 불가'(=AXXX00)"},
     # 고객사 라벨은 여기에 적지 않는다 — 실명이 공개 저장소에 남기 때문이다.
     # 마스킹 맵(sources/mask_map.json)이 있으면 별칭으로, 없으면 라벨 없이(코드 그대로) 간다.
     # derive_customer_labels()가 룰 메시지에서 코드↔이름 근거를 찾아 채운다.
     {"field": "PAK_MSG_CD", "code": "203", "label": "별도기준적용+SKID변경적용",
      "certainty": "추정", "evidence": "order#16 '별도기준적용+SKID변경적용 선택 필수'(포장메세지 203)"}]
    + [{"field": "EMBS_CD", "code": c, "label": "양면엠보",
        "certainty": "추정", "evidence": "order#85 '양면엠보제품 지관삽입필수'(EMBOSS 2L·2S·2W)"}
       for c in ("2L", "2S", "2W")]
    + [{"field": "ORD_USG_CD", "code": f"{ch}ZZ000", "label": "이행용",
        "certainty": "추정", "evidence": "order#10 '이행용 용도코드(사용불가)'(11개 코드 목록)"}
       for ch in "ABCDEFGHIJK"]
)

# 축소 드리프트 5건 (M5, needs_review) — order 행번호 기준
REDUCTION_ORDER_ROWS = {
    73: "A71 조도 필수 품명에서 7 제거 (cross#61은 7 포함) — 의도적 완화 vs 미반영 확인 필요",
    75: "A72 GI Spangle 금지값 {1,6}→{6}·품명 {G}→{G,K,V} — 검사 값 자체가 다른 유일한 Spangle 쌍",
    84: "A732 EMBOSS 불필요 품명에서 C 제거 (cross#72는 C 포함)",
    88: "A671 보호필름폭 불필요 품명에서 E,G,L,J 제거 (cross#76이 4종 더 넓음)",
    95: "A441 최소단중 임계 2000→1 (룰 사실상 무력화)·품명 +K,V,W — 오입력 여부 확인 필수",
}

CATEGORY_ORDER = ["누락", "사용불가", "불필요", "한도위반", "선택오류"]


class PipelineError(Exception):
    pass


def _cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


# ---------------------------------------------------------------------------
# Task 1 Step 1: 파서 (C-1, C-2)
# ---------------------------------------------------------------------------
def parse_workbook(path):
    """xlsx -> {meta, fields, rules}. 룰 인정 C-1(no.가 ^\\d+$), 조건 인정 C-2(NOT_CHECK/None 제외)."""
    if not Path(path).exists():
        raise PipelineError(
            f"입력 파일 없음: {path}\n"
            f"  원본 xlsx 2종({ORDER_XLSX_NAME} · {CROSS_XLSX_NAME})은 저장소에 포함되지 않습니다.\n"
            f"  {SOURCE_DIR}/ 에 두거나, OC_XLSX_DIR=<디렉토리> 환경변수 또는\n"
            f"  --order <경로> --cross <경로> 로 지정하세요 (build/README.md '작업 순서' 0번)."
        )
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    meta = {
        "기준코드": _cell(ws, 2, 1), "버전": _cell(ws, 2, 2), "유형": _cell(ws, 2, 3),
        "상태": _cell(ws, 2, 4), "시작일자": _cell(ws, 2, 5), "기준설명": _cell(ws, 2, 6),
    }
    groups = []  # (한국어 필드명, 시작 컬럼) — 연산자/비교값1/비교값2 3컬럼 묶음
    err_code_col = err_msg_col = None
    for c in range(1, ws.max_column + 1):
        name = _cell(ws, 4, c)
        if not name:
            continue
        if name == "주문에러코드":
            err_code_col = c
        elif name == "주문에러메세지":
            err_msg_col = c
        else:
            groups.append((name, c))
    if err_code_col is None or err_msg_col is None:
        raise PipelineError(f"{path}: 에러코드/메세지 컬럼을 찾지 못함")
    rules = []
    for r in range(6, ws.max_row + 1):
        no = _cell(ws, r, 1)
        if no is None or not re.fullmatch(r"\d+", no):
            continue  # C-1: 범례 END!/※/빈 행 제거
        conds = []
        for name, c in groups:
            op = _cell(ws, r, c)
            if op is None or op == "NOT_CHECK":
                continue  # C-2
            conds.append({"field": name, "op": op,
                          "v1": _cell(ws, r, c + 1), "v2": _cell(ws, r, c + 2)})
        if not conds:
            raise PipelineError(f"{path} 행 {no}: 조건 0개 룰 — 로드 거부 (C-2)")
        rules.append({"no": int(no), "conds": conds,
                      "err_code": _cell(ws, r, err_code_col),
                      "err_msg": _cell(ws, r, err_msg_col) or ""})
    return {"meta": meta, "fields": [g[0] for g in groups], "rules": rules}


# ---------------------------------------------------------------------------
# C-3 / C-4: 메시지 클렌징
# ---------------------------------------------------------------------------
TAG_RE = re.compile(r"^\[(?P<dept>[^\]]+)\]\[(?P<actor>[^\]]+)\]\s*")

CROSS_TEST_FIXES = {  # C-4: (행번호) -> (원문 기대값, 정정값)
    1: ("주문요청번호 누락,테스트1", "주문요청번호 누락"),
    2: ("주문요청행번 누락1", "주문요청행번 누락"),
}


def clean_message(raw_msg, source, row):
    """C-3 태그 분해 + C-4 테스트 흔적 제거 + 끝 콤마 제거. -> (dept, actor, body, flags)"""
    msg = raw_msg
    if source == "cross" and row in CROSS_TEST_FIXES:
        expect, fixed = CROSS_TEST_FIXES[row]
        if msg != expect:
            raise PipelineError(f"C-4 전제 불일치: cross#{row} 원문 {msg!r} != {expect!r}")
        msg = fixed
    dept = actor = None
    m = TAG_RE.match(msg)
    if m:
        dept, actor = m.group("dept"), m.group("actor")
        msg = msg[m.end():]
    msg = msg.strip()
    flags = []
    if msg.endswith(","):
        msg = msg.rstrip(",").strip()
        flags.append("truncated_message")  # order#135/#136 (아래에서 어서션)
    return dept, actor, msg, flags


# ---------------------------------------------------------------------------
# 필드 사전 구축 (라벨 도출 -> 타입 판정)
# ---------------------------------------------------------------------------
LABEL_PATTERNS = [
    re.compile(r"^(.+?)\s*누락$"),
    re.compile(r"^(.+?)\s*불필요$"),
    re.compile(r"^(.+?)\s*오류(\(.*\))?$"),
]


def derive_labels(parsed_by_source):
    """EAV 타깃 코드 -> 라벨. 전용 컬럼 매핑 라벨 우선, 그 외에는 누락/불필요/오류 메시지에서 도출."""
    labels = dict(KEY_COLUMN_LABEL)  # 플랜 매핑 라벨이 우선
    for source, parsed in parsed_by_source.items():
        for rule in parsed["rules"]:
            item = next((c["v1"] for c in rule["conds"] if c["field"] == "주문항목코드"), None)
            if item is None or item in labels:
                continue
            _, _, body, _ = clean_message(rule["err_msg"], source, rule["no"])
            for pat_i, pat in enumerate(LABEL_PATTERNS):
                m = pat.match(body)
                if m:
                    labels[item] = m.group(1).strip()
                    break
    # 두 번째 패스: 우선순위 높은 패턴(누락)이 다른 행에 있으면 그걸 채택
    for source, parsed in parsed_by_source.items():
        for rule in parsed["rules"]:
            item = next((c["v1"] for c in rule["conds"] if c["field"] == "주문항목코드"), None)
            if item is None or item in KEY_COLUMN_LABEL:
                continue
            _, _, body, _ = clean_message(rule["err_msg"], source, rule["no"])
            m = LABEL_PATTERNS[0].match(body)
            if m:
                labels[item] = m.group(1).strip()
    return labels


def field_type(key, label):
    if label and any(k in label for k in NUMBER_KEYWORDS) \
            and not label.endswith(NUMBER_EXCLUDE_SUFFIX):
        return "number"
    if label and label.endswith(CODE_LABEL_SUFFIX):
        return "code"
    if key.endswith(CODE_KEY_SUFFIX):
        return "code"
    return "string"


# ---------------------------------------------------------------------------
# Task 1 Step 2~3: 조건 클렌징 (C-5/C-6/C-7) + EAV 승격
# ---------------------------------------------------------------------------
def _num(v, ctx):
    try:
        f = float(v)
    except (TypeError, ValueError):
        raise PipelineError(f"{ctx}: 숫자 연산자에 비숫자 값 {v!r} (C-5 위반)")
    return int(f) if f == int(f) else f


def _display_hint(op, num):
    """C-6 ±0.001/+1 보정값 표시용 자연어. 소수 3자리&끝자리1(LT/GE), 정수&끝자리1(GE, >=100)."""
    s = str(num)
    if op == "LT" and re.fullmatch(r"\d+\.\d{2}1", s):
        return f"{round(num - 0.001, 3):g} 이하"
    if op == "GE" and re.fullmatch(r"\d+\.\d{2}1", s):
        return f"{round(num - 0.001, 3):g} 초과"
    if op == "GE" and isinstance(num, int) and num >= 100 and s.endswith("1"):
        return f"{num - 1:g} 초과"
    return None


def clean_condition(field_key, op, v1, v2, ftype, ctx, stats):
    """단일 조건 정규화 -> dict. 숫자는 float 승격(문자열 비교 금지), BETWEEN1=[l,h] / BETWEEN2=[l,h)."""
    if op in NULL_OPS:
        if v1 is not None or v2 is not None:
            raise PipelineError(f"{ctx}: {op}에 비교값 존재 (정합성 위반)")
        return {"field": field_key, "op": op}
    if op in ("BETWEEN1", "BETWEEN2"):
        low, high = _num(v1, ctx), _num(v2, ctx)
        if low > high:
            raise PipelineError(f"{ctx}: BETWEEN 값1({low}) > 값2({high})")
        return {"field": field_key, "op": "BETWEEN", "low": low, "high": high,
                "lowInclusive": True, "highInclusive": op == "BETWEEN1"}
    if op in (">", ">=", "<", "<="):
        if v2 is not None:
            raise PipelineError(f"{ctx}: {op}에 비교값2 존재")
        num = _num(v1, ctx)
        cond = {"field": field_key, "op": OP_MAP[op], "values": [num]}
        hint = _display_hint(cond["op"], num)
        if hint:
            cond["display"] = hint
            stats["display_hints"].append({"where": ctx, "op": cond["op"], "value": num, "display": hint})
        return cond
    if op in ("IN", "NOT_IN"):
        if v1 is None:
            raise PipelineError(f"{ctx}: {op}에 값 없음")
        items = [x.strip() for x in str(v1).split(",")]
        items = sorted(set(x for x in items if x))  # C-7: trim → 빈 원소 제거 → 정렬 → dedupe
        if not items:
            raise PipelineError(f"{ctx}: {op} 원소 0개")
        if ftype == "number":
            items = sorted(_num(x, ctx) for x in items)
        if len(items) == 1:  # C-7: 단일 원소는 EQ/NE로 강등
            stats["in_demoted" if op == "IN" else "not_in_demoted"] += 1
            return {"field": field_key, "op": "EQ" if op == "IN" else "NE", "values": items}
        return {"field": field_key, "op": op, "values": items}
    if op in ("=", "!="):
        if v2 is not None:
            raise PipelineError(f"{ctx}: {op}에 비교값2 존재")
        if v1 is None:
            raise PipelineError(f"{ctx}: {op}에 비교값 없음")
        val = _num(v1, ctx) if ftype == "number" else str(v1)
        return {"field": field_key, "op": OP_MAP[op], "values": [val]}
    raise PipelineError(f"{ctx}: 미지원 연산자 {op!r}")


def clean_rules(parsed, source, labels, stats):
    """clean_rules(raw) -> [CleanRule]. EAV 승격 포함 (Task 1 Step 3)."""
    out = []
    for rule in parsed["rules"]:
        row = rule["no"]
        ctx0 = f"{source}#{row}"
        dept, actor, body, flags = clean_message(rule["err_msg"], source, row)
        item = None
        value_conds, other_conds = [], []
        for c in rule["conds"]:
            if c["field"] == "주문항목코드":
                if c["op"] != "=" or not c["v1"]:
                    raise PipelineError(f"{ctx0}: 주문항목코드 조건이 '=' 단일값이 아님: {c}")
                if item is not None:
                    raise PipelineError(f"{ctx0}: 주문항목코드 조건 중복 (multi-EAV)")
                item = c["v1"]
            elif c["field"] == "주문항목값":
                value_conds.append(c)
            else:
                other_conds.append(c)
        when, context = [], []
        if item is not None:
            if not value_conds:
                raise PipelineError(f"{ctx0}: EAV 앵커({item})만 있고 주문항목값 조건 없음")
            ftype = field_type(item, labels.get(item))
            for c in value_conds:
                when.append(clean_condition(item, c["op"], c["v1"], c["v2"], ftype,
                                            f"{ctx0} 주문항목값({item})", stats))
        elif value_conds:
            raise PipelineError(f"{ctx0}: 주문항목값 조건이 있으나 주문항목코드 앵커 없음")
        for c in other_conds:
            key = COLUMN_KEY_MAP.get(c["field"])
            if key is None:
                raise PipelineError(f"{ctx0}: 매핑 없는 전용 컬럼 {c['field']!r}")
            context.append(clean_condition(key, c["op"], c["v1"], c["v2"],
                                           field_type(key, labels.get(key)),
                                           f"{ctx0} {c['field']}", stats))
        if item is None:
            # EAV 앵커 없는 룰: target은 when의 첫 필드로 (현 데이터에는 없음 — 방어 코드)
            when, context = context[:1], context[1:]
            item = when[0]["field"]
        out.append({
            "source": source, "row": row, "err_code": rule["err_code"],
            "target": {"item": item, "label": labels.get(item, item)},
            "when": when, "context": context,
            "message": body, "owner_dept": dept, "actor": actor, "flags": flags,
            "raw_conds": rule["conds"], "raw_message": rule["err_msg"],
        })
    return out


# ---------------------------------------------------------------------------
# 병합 (Task 2 Step 1~3, M1~M9)
# ---------------------------------------------------------------------------
def cond_sig(c):
    if c["op"] == "BETWEEN":
        return (c["field"], "BETWEEN", c["low"], c["high"], c["lowInclusive"], c["highInclusive"])
    if c["op"] in NULL_OPS:
        return (c["field"], c["op"])
    return (c["field"], c["op"], tuple(c["values"]))


def rule_sig(r):
    return frozenset(cond_sig(c) for c in r["when"] + r["context"])


def conds_json(conds):
    return [dict(c) for c in conds]


def lineage_entry(r, diff=None):
    e = {"source": r["source"], "row": r["row"], "err_code": r["err_code"],
         "raw_message": r["raw_message"], "conds": r["raw_conds"]}
    if diff:
        e["diff"] = diff
    return e


def diff_conditions(order_rule, cross_rule, labels):
    """두 룰의 when+context를 필드별로 대조해 사람이 읽는 diff 문자열 생성."""
    def by_field(conds):
        d = {}
        for c in conds:
            d.setdefault(c["field"], []).append(c)
        return d
    o, c = by_field(order_rule["when"] + order_rule["context"]), \
        by_field(cross_rule["when"] + cross_rule["context"])
    parts = []

    def fmt(cs):
        out = []
        for x in cs:
            if x["op"] == "BETWEEN":
                out.append(f"BETWEEN {x['low']}~{x['high']}")
            elif x["op"] in NULL_OPS:
                out.append(x["op"])
            else:
                out.append(f"{x['op']} {','.join(str(v) for v in x['values'])}")
        return " & ".join(out)
    for f in sorted(set(o) | set(c)):
        so, sc = fmt(o.get(f, [])), fmt(c.get(f, []))
        if so != sc:
            parts.append(f"{labels.get(f, f)}: cross[{sc or '없음'}] → order[{so or '없음'}]")
    if order_rule["err_code"] != cross_rule["err_code"]:
        parts.append(f"err_code: cross {cross_rule['err_code']} → order {order_rule['err_code']}")
    if order_rule["message"] != cross_rule["message"]:
        parts.append("message 상이")
    return " · ".join(parts) if parts else None


def build_ruleset(order_clean, cross_clean, labels, stats):
    order_by_row = {r["row"]: r for r in order_clean}
    # --- M1: 조건 시그니처 병합 ---
    order_by_sig = {}
    for r in order_clean:
        s = rule_sig(r)
        if s in order_by_sig:
            raise PipelineError(f"order 내부 시그니처 중복: #{order_by_sig[s]['row']} vs #{r['row']}")
        order_by_sig[s] = r
    pairs, cross_rest = [], []
    for r in cross_clean:
        s = rule_sig(r)
        if s in order_by_sig:
            pairs.append((order_by_sig[s], r))
        else:
            cross_rest.append(r)
    stats["sig_pairs"] = len(pairs)
    if len(pairs) != 78:
        raise PipelineError(f"조건 동일 쌍 {len(pairs)}건 — 기대 78 (M1)")

    merged = {}   # order row -> rule record(작업 중)
    for r in order_clean:
        merged[r["row"]] = {
            "clean": r, "cross_lineage": [], "merge_status": "auto",
            "merge_rule": "M9", "merge_reason": "order 고유 룰 — 그대로 탑재",
            "legacy_codes": [], "review_alternatives": None,
            "source_tag": "order", "flags": list(r["flags"]), "enabled": True,
        }

    redirect_done = False
    for o, c in pairs:
        if o["row"] == 77 and c["row"] == 69:
            # M3: cross#69(A722, 품명 3,6)는 order#81(A722, 품명 3,6,9)의 부분집합 — #81로 흡수
            t = merged[81]
            t["cross_lineage"].append(lineage_entry(
                c, diff="M3 흡수: A722 부분집합(품명 3,6 ⊆ 3,6,9) — order#77(A72)와 코드·메시지 상충했던 쌍"))
            t["merge_rule"] = "M3"
            t["merge_reason"] = ("cross#69(A722 '컬러원판 EXTRA SMOOTH ONLY')는 본 룰의 부분집합이라 흡수 — "
                                 "조건 동일이 아니다(품명 cross[3,6] ⊂ order[3,6,9]). 그래서 출처는 '공통'이 "
                                 "아니라 'order'. order#77(A72 CCGI)과 품명 3,6에서 동시발화가 설계에 내재 — 상충 아님")
            # source_tag는 'order' 유지 — '공통'은 조건 시그니처 완전 동일만 (라운드1 fidelity 지적)
            if "subset_absorbed" not in t["flags"]:
                t["flags"].append("subset_absorbed")
            redirect_done = True
            continue
        t = merged[o["row"]]
        diff = diff_conditions(o, c, labels)
        t["cross_lineage"].append(lineage_entry(c, diff=diff))
        t["source_tag"] = "공통"
        t["merge_rule"] = "M1"
        t["merge_reason"] = "조건 완전 동일 — order 메시지·코드 채택, cross는 계보 보존"
        if o["row"] == 10 and c["row"] == 8:
            # M2: 에러코드 상충 A893(order) vs A89(cross) — order 채택 + legacy 보존
            t["merge_rule"] = "M2"
            t["merge_status"] = "conflict_resolved"
            t["legacy_codes"] = ["A89"]
            t["merge_reason"] = ("에러코드 상충: cross#8 A89 vs order#10 A893. "
                                 "누락(A89)과 사용불가를 분리한 세분화로 판단, order 코드 채택 + legacy_codes 보존")
    if not redirect_done:
        raise PipelineError("M3 전제 실패: (order#77, cross#69) 시그니처 쌍이 없음")
    for row in (77, 81, 87, 115):  # 그림자/동시발화 쌍 (C-11, M3)
        if "co_fire_pair" not in merged[row]["flags"]:
            merged[row]["flags"].append("co_fire_pair")

    # --- 특례 분리: cross#50(A321, M7) / cross#98(A92 IS_NULL, M8) / cross#85(A451, M6 이슈 전용) ---
    specials = {}
    rest2 = []
    for r in cross_rest:
        if r["row"] in (50, 85, 98):
            specials[r["row"]] = r
        else:
            rest2.append(r)
    for need in (50, 85, 98):
        if need not in specials:
            raise PipelineError(f"cross#{need} 특례 룰이 미매칭 잔여에 없음")

    # --- 드리프트 (M4/M5): (target.item, err_code, message) 매칭 ---
    matched_order_rows = {o["row"] for o, c in pairs if not (o["row"] == 77 and c["row"] == 69)} | {81}
    pool = {}
    for r in order_clean:
        if r["row"] in matched_order_rows:
            continue
        key = (r["target"]["item"], r["err_code"], r["message"])
        pool.setdefault(key, []).append(r)
    expansion, reduction = [], []
    for c in rest2:
        key = (c["target"]["item"], c["err_code"], c["message"])
        cands = pool.get(key, [])
        if len(cands) != 1:
            raise PipelineError(
                f"cross#{c['row']} 드리프트 상대 {'없음' if not cands else '모호'}"
                f"({[x['row'] for x in cands]}): {key}")
        o = cands[0]
        t = merged[o["row"]]
        diff = diff_conditions(o, c, labels)
        if o["row"] in REDUCTION_ORDER_ROWS:
            # M5: 범위 축소 — order 값 탑재 + needs_review + cross 대안 병기
            t["merge_rule"] = "M5"
            t["merge_status"] = "needs_review"
            t["merge_reason"] = f"범위 축소 드리프트: {REDUCTION_ORDER_ROWS[o['row']]}"
            t["review_alternatives"] = [{
                "source": "cross", "row": c["row"], "err_code": c["err_code"],
                "when": conds_json(c["when"]), "context": conds_json(c["context"]),
            }]
            t["cross_lineage"].append(lineage_entry(c, diff=f"M5 축소 · {diff}"))
            reduction.append((o, c))
        else:
            # M4: 범위 확장 — order 채택 auto, cross 범위는 lineage.diff
            t["merge_rule"] = "M4"
            t["merge_reason"] = "범위 확장 드리프트: order가 신품명(N,K,V,W,8,9 등) 반영 확장본 — order 채택"
            t["cross_lineage"].append(lineage_entry(c, diff=f"M4 확장 · {diff}"))
            expansion.append((o, c))
    if len(reduction) != 5:
        raise PipelineError(f"축소 드리프트 {len(reduction)}건 — 기대 5 (M5)")
    if len(expansion) != 12:
        raise PipelineError(f"확장 드리프트 {len(expansion)}건 — 기대 12 (M4)")
    if {o["row"] for o, _ in reduction} != set(REDUCTION_ORDER_ROWS):
        raise PipelineError("축소 드리프트 대상 행 불일치")

    # --- M6: order A451 가족 채택, cross#85(>8000)는 룰 미탑재 (union 금지) ---
    # 폐기하되 계보는 남긴다 — expert2 §5 M6 완화책("cross#85를 lineage에 남겨 구기준 8000 대조 표시").
    c85 = specials[85]
    for row in (98,):  # order#98 = A451 재정의 본체(동일 코드 A451·동일 대상 항목)
        merged[row]["merge_rule"] = "M6"
        merged[row]["merge_reason"] = ("A451 가족 재설계: cross#85(>8000 일반 상한)는 룰로 탑재하지 않고 "
                                       "본 룰 lineage + 이슈 I-C3으로만 기록. "
                                       "union 시 신정책이 허용한 주문을 구정책이 기각하는 역행 발생")
        merged[row]["cross_lineage"].append(lineage_entry(
            c85, diff="M6 폐기: 구 일반상한 >8000(품명 1~6·제품형태 C) — 룰 미탑재. "
                      "A451~A460 두께·수요가별 세분 가족(5000~16000)으로 대체(union 금지 · 이슈 I-C3)"))

    # --- M3 보칙: 시그니처 완전 동일 쌍의 실체는 (order#77, cross#69)였다 ---
    # 코드·메시지(A72 'CCGI Spangle 선택 오류')로 재짝지어 cross#65를 order#77의 계보로 채택했으므로,
    # order#77의 출처는 '공통'이 아니라 'order'(M4 확장)다. 재짝짓기 사실을 룰에 명시한다.
    merged[77]["merge_reason"] += (
        " · 재짝짓기 주의: 조건 시그니처가 완전히 동일한 cross 행은 cross#65가 아니라 cross#69(A722·품명 3,6)였다. "
        "코드·메시지가 일치하는 cross#65(A72·품명 3)를 계보로 채택하고 cross#69는 order#81로 흡수했다(M3·이슈 I-C2). "
        "따라서 78쌍 중 이 1쌍만 '공통' 배지를 받지 않는다")

    stats["drift_expansion"] = len(expansion)
    stats["drift_reduction"] = len(reduction)

    # --- 최종 룰 배열 조립 ---
    rules = []
    for row in sorted(merged):
        t = merged[row]
        r = t["clean"]
        error = {"code": r["err_code"], "legacy_codes": t["legacy_codes"],
                 "message": r["message"], "category": categorize(r["message"]),
                 "severity": "error"}
        wc = warn_candidate(r["message"])
        if wc:
            error["severity_candidate"] = "warn"
        error["owner_dept"] = r["owner_dept"]
        error["actor"] = r["actor"]
        merge = {"status": t["merge_status"], "rule": t["merge_rule"], "reason": t["merge_reason"]}
        if t["review_alternatives"]:
            merge["review_alternatives"] = t["review_alternatives"]
        rules.append({
            "id": f"R-{row:03d}", "enabled": t["enabled"], "source": t["source_tag"],
            "target": dict(r["target"]),
            "when": conds_json(r["when"]), "context": conds_json(r["context"]),
            "error": error, "merge": merge, "flags": t["flags"],
            "lineage": [lineage_entry(r)] + t["cross_lineage"],
        })

    # M8: cross#98 — order#112(=0)와 상보적 IS_NULL 변형, 별도 룰로 탑재
    c98 = specials[98]
    rules.append(make_cross_rule(
        c98, enabled=True, status="auto", mrule="M8",
        reason="order#112(조합폭10 =0)와 상보: NULL 케이스와 0 케이스를 각각 검출. 채택 비용 0·방어 이득 양수"))
    # M7: cross#50 A321 — 기본 비활성 + needs_review
    c50 = specials[50]
    rules.append(make_cross_rule(
        c50, enabled=False, status="needs_review", mrule="M7",
        reason="cross에만 존재(order에는 항목코드 ORD_SHT_CNT 자체가 없음). "
               "의도적 폐지 vs 이관 누락 판단 필요 — 기본 비활성"))

    return rules, specials[85], {"pairs": pairs, "expansion": expansion, "reduction": reduction}


def make_cross_rule(r, enabled, status, mrule, reason):
    error = {"code": r["err_code"], "legacy_codes": [], "message": r["message"],
             "category": categorize(r["message"]), "severity": "error",
             "owner_dept": r["owner_dept"], "actor": r["actor"]}
    if warn_candidate(r["message"]):
        error["severity_candidate"] = "warn"
    return {
        "id": f"R-C{r['row']:02d}", "enabled": enabled, "source": "cross",
        "target": dict(r["target"]),
        "when": conds_json(r["when"]), "context": conds_json(r["context"]),
        "error": error,
        "merge": {"status": status, "rule": mrule, "reason": reason},
        "flags": list(r["flags"]), "lineage": [lineage_entry(r)],
    }


# ---------------------------------------------------------------------------
# Task 2 Step 4: category / severity / 이슈 / 사전
# ---------------------------------------------------------------------------
def categorize(msg):
    body = msg.rstrip(")").rstrip()
    if body.endswith("누락"):
        return "누락"
    if body.endswith("불필요"):
        return "불필요"
    if body.endswith("불가") or body.endswith("사용불가"):
        return "사용불가"
    if any(k in msg for k in ("한도", "준수", "제한", "Max", "Min", "MAX", "MIN")):
        return "한도위반"
    return "선택오류"


def warn_candidate(msg):
    return any(k in msg for k in ("협의", "문의", "확인"))


def build_issues(rules, cross85, detail, labels):
    by_id = {r["id"]: r for r in rules}

    def snap(r):
        return {"source": r["source"], "row": r["row"], "err_code": r["err_code"],
                "message": r["message"],
                "when": conds_json(r["when"]), "context": conds_json(r["context"])}
    issues = []
    # --- 상충 2건 ---
    o10 = next(o for o, c in detail["pairs"] if o["row"] == 10)
    c8 = next(c for o, c in detail["pairs"] if o["row"] == 10)
    issues.append({
        "id": "I-C1", "type": "conflict", "title": "에러코드 상충: A89(cross) vs A893(order) — 이행용 용도코드",
        "body": "조건이 완전히 동일한데 에러코드가 다르다. order가 누락(A89)과 사용불가(A893)를 분리한 "
                "세분화로 판단해 order 코드를 채택했고, 구코드는 legacy_codes로 보존했다.",
        "resolution": "order 채택으로 해소됨 (R-010, legacy_codes:[A89])",
        "rule_ids": ["R-010"], "left": snap(o10), "right": snap(c8), "status": "open",
    })
    o77r = by_id["R-077"]
    c69 = next(c for o, c in detail["pairs"] if c["row"] == 69 and c["source"] == "cross")
    issues.append({
        "id": "I-C2", "type": "conflict", "title": "코드·메시지 상충: order#77 A72 'CCGI Spangle 선택 오류' vs cross#69 A722 '컬러원판 EXTRA SMOOTH ONLY'",
        "body": "같은 조건(값 1,3,6·품명 3,6)에 코드도 메시지도 다르다. cross#69는 order#81(A722, 품명 3,6,9)의 "
                "부분집합이라 order#81의 계보로 흡수했고, order#77(A72)은 별도 유지했다. 품명 3,6에서 두 룰의 "
                "동시발화는 설계에 내재된 다중 결함 안내다(상충 아님).",
        "resolution": "cross#69 → R-081 흡수, R-077 별도 유지, 둘 다 co_fire_pair 플래그",
        "rule_ids": ["R-077", "R-081"],
        "left": {"source": "order", "row": 77, "err_code": "A72", "message": o77r["error"]["message"],
                 "when": o77r["when"], "context": o77r["context"]},
        "right": snap(c69),
        "status": "open",
    })
    # --- 상충 3건째: A451 정책 재설계 (cross#85 폐기) ---
    # 룰로는 탑재하지 않지만(union 금지) 원본 98행 중 유일하게 계보가 소실되던 행이라
    # 조건 전문을 이슈에 보존한다 — expert2 §5 M6 완화책 + 플랜 Task 2 Step 3.
    fam = [r for r in rules
           if re.fullmatch(r"A4(5[1-9]|60)", r["error"]["code"] or "")]
    r98 = by_id["R-098"]
    issues.append({
        "id": "I-C3", "type": "conflict",
        "title": "정책 재설계 상충: cross#85 A451 일반 상한(>8000) vs order A451~A460 두께·수요가별 세분 가족",
        "body": "cross는 '컬러 코일형 최대단중 준수'를 품명 1~6·제품형태 C에 대해 단중상한 8000 초과 일괄 금지 "
                "1룰로 걸었다. order는 같은 코드 A451을 두께·폭·보호필름·최종수요가별 세분 가족(A451~A460, "
                f"임계 5000~16000, 총 {len(fam)}룰)으로 재설계했다. 두 기준을 union하면 신정책이 허용한 주문을 "
                "구정책이 기각한다 — 검증 반례: 품명 2·두께 1.2·단중상한 12000 주문은 order A452(>16000)로는 "
                "통과지만 cross 구룰(>8000)로는 기각. 따라서 cross#85는 룰로 탑재하지 않는다. "
                "다만 order 세분 가족에 8000 일반 상한 구간이 실제로 누락됐는지(예: 0.9T 컬러 코일 12톤 주문)는 "
                "현업 확인 대상이므로 구기준 조건 전문을 여기 보존한다.",
        "resolution": "order 가족 채택 · cross#85 룰 미탑재(union 금지). 계보는 R-098 lineage + 본 이슈에 보존",
        "rule_ids": [r["id"] for r in fam],
        "left": {"source": "order", "row": 98, "err_code": r98["error"]["code"],
                 "message": r98["error"]["message"],
                 "when": r98["when"], "context": r98["context"]},
        "right": snap(cross85),
        "status": "open",
    })
    # --- 확인필요 6건 ---
    review_meta = {
        "R-095": ("A441 최소단중 임계값: order 1kg vs cross 2000kg",
                  "order 기준으로는 하한 1kg — 2톤 하한 검증이 사실상 무력화된 상태. 의도된 완화(소단중 허용)인지 "
                  "오입력(1000/2000 탈자)인지 원천 확인 필수. 단위 kg은 order#135의 12500(=12.5톤)으로 확정."),
        "R-073": ("A71 조도코드 필수 품명: order는 7 제외 vs cross는 7 포함",
                  "품명 7이 order에선 '선택', cross에선 '필수'로 정면 모순. order#74 불필요 NOT_IN에는 7이 남아 있어 "
                  "order 내부만 보면 '선택으로 완화'라는 일관 해석 가능 — 확인 전까지 모순 상태."),
        "R-084": ("A732 EMBOSS 불필요 품명: order A,B,D vs cross A,B,C,D",
                  "cross에만 품명 C 포함. order의 의도적 축소인지 편집 누락인지 확인 필요."),
        "R-088": ("A671 보호필름폭 불필요 품명: order A,B,C,D vs cross A,B,C,D,E,G,L,J",
                  "cross가 E,G,L,J 4종 더 넓다. 방향이 '축소'인 드리프트로 자동 채택하지 않음."),
        "R-075": ("A72 GI Spangle 금지값: order {6}(품명 G,K,V) vs cross {1,6}(품명 G)",
                  "검사 값 집합 자체가 다른 유일한 Spangle 쌍. 값 '1'이 order에서는 허용, cross에서는 에러."),
        "R-C50": ("A321 주문Sheet매수 불필요(cross 고유) — 기본 비활성",
                  "order에는 대응 룰도 항목코드도 없음. 탑재 시 폐지된 제약 부활 위험, 미탑재 시 이관 누락 가능성(경미). "
                  "위험 비대칭상 기본 비활성으로 탑재하고 판단은 현업 몫."),
    }
    for rid, (title, body) in review_meta.items():
        r = by_id[rid]
        issues.append({
            "id": f"I-R{len([i for i in issues if i['type'] == 'needs_review']) + 1}",
            "type": "needs_review", "title": title, "body": body,
            "rule_ids": [rid],
            "order_value": {"when": r["when"], "context": r["context"], "enabled": r["enabled"]},
            "cross_alternative": (r["merge"].get("review_alternatives") or [None])[0],
            "status": "open",
        })
    # --- 데이터 결함 5건 ---
    issues.append({
        "id": "I-D1", "type": "data_defect", "title": "조합폭 누락 체크의 NULL 구멍",
        "body": "order#103~112는 조합폭N '= 0'일 때만 누락 판정하고 IS_NULL 변형이 없다. 값이 미입력(NULL)이면 "
                "누락이 통과된다. 단중·중량 계열(A44/A45/A51/A27/A69)은 =0과 IS_NULL 두 룰을 모두 두는 관례를 "
                "지켰는데 조합폭만 누락. cross#98(조합폭10 IS_NULL)만 예외적으로 존재해 R-C98로 탑재했다. "
                "엔진 보정 없이 결함으로만 노출(룰 발명 금지).",
        "rule_ids": [f"R-{i:03d}" for i in range(103, 113)] + ["R-C98"], "status": "open",
    })
    issues.append({
        "id": "I-D2", "type": "data_defect", "title": "A454 3룰이 동일 코드·동일 메시지에 임계값만 다름",
        "body": "폭 [800,1101) 2.2톤 / [1101,1301) 2.4톤 / [1301,1601) 2.6톤인데 메시지가 셋 다 같아 "
                "에러를 받은 사용자가 자기 한도를 알 수 없다. 폭 <800, >=1601 구간은 체크 자체가 없다(의도 여부 확인 필요).",
        "rule_ids": ["R-132", "R-133", "R-134"], "status": "open",
    })
    issues.append({
        "id": "I-D3", "type": "data_defect", "title": "메시지 잘림 의심 2건 (끝 콤마)",
        "body": "order#135/#136 메시지가 콤마로 끝난다('…Max 12.5톤,' / '…Max 8.5톤,'). 뒷부분 잘림 의심 — "
                "원문은 lineage.raw_message에 보존, truncated_message 플래그 부착.",
        "rule_ids": ["R-135", "R-136"], "status": "open",
    })
    issues.append({
        "id": "I-D4", "type": "data_defect", "title": "의미론적 포함(그림자) 룰 3쌍 — 동시발화 안내",
        "body": "order#77[A72] ⊆ order#81[A722], cross#65[A72] ⊆ cross#69[A722](병합 후 R-077/R-081로 수렴), "
                "order#115[A664] ⊆ order#87[A663]. 좁은 룰이 발화하는 모든 주문에서 넓은 룰도 발화한다 — "
                "모순이 아니라 다중 결함이며, first-match 엔진이면 한쪽이 영구 은폐된다(전수 평가 채택 근거).",
        "rule_ids": ["R-077", "R-081", "R-087", "R-115"], "status": "open",
    })
    issues.append({
        "id": "I-D5", "type": "data_defect", "title": "에러코드 네이밍 충돌: A46 vs A460 (+A894 결번)",
        "body": "A46(제품형태 누락)과 A460(A45 단중 가족의 연속 번호)이 별개 코드로 공존 — 접두사 기반 "
                "그룹핑(A46*)을 하면 남의 코드가 섞인다. A89 계열은 A891,A892,A893,A895로 A894 결번.",
        "rule_ids": ["R-004", "R-136"], "status": "open",
    })
    return issues


# ---------------------------------------------------------------------------
# 필드 사전 / 시드 조립
# ---------------------------------------------------------------------------
def build_fields(rules, labels, parsed_order, parsed_cross):
    used = {}  # key -> {"eav": bool, "column": bool, "refs": int}
    def touch(key, origin):
        e = used.setdefault(key, {"eav": False, "column": False, "refs": 0})
        e[origin] = True
        e["refs"] += 1
    for r in rules:
        touch(r["target"]["item"], "eav")
        used[r["target"]["item"]]["refs"] -= 1  # target 자체는 when에서 다시 세므로 중복 보정
        for c in r["when"]:
            touch(c["field"], "eav")
        for c in r["context"]:
            touch(c["field"], "column")
        for alt in (r["merge"].get("review_alternatives") or []):
            for c in alt["when"] + alt["context"]:
                used.setdefault(c["field"], {"eav": False, "column": False, "refs": 0})
    fields = []
    for key in sorted(used):
        label = labels.get(key)
        if label is None:
            if key not in LABEL_UNKNOWN_OK:
                raise PipelineError(f"필드 {key} 라벨 도출 실패 (LABEL_UNKNOWN_OK 미등재)")
            label = key
        origin = ("EAV|column" if used[key]["eav"] and used[key]["column"]
                  else "column" if used[key]["column"] else "EAV")
        f = {"key": key, "label": label, "type": field_type(key, labels.get(key)),
             "origin": origin, "refs": used[key]["refs"]}
        if key in FIELD_UNITS:
            f["unit"] = FIELD_UNITS[key]
        if key in LABEL_UNKNOWN_OK:
            f["label_note"] = "라벨 미상 — 원문 코드 그대로(메시지에서 역추론 불가, 날조 금지)"
        if key == "PAK_MSG_CD":
            f["label_note"] = ("임시 영문키(원본 EAV 근거 없음). 플랜의 ORD_TEM_CD 매핑은 "
                               "EAV 영업팀코드(A18)와 상충하여 회피 — pipeline_notes.md D-1")
        fields.append(f)
    return fields


#: 공개 배포용 마스킹 — 저장소가 public이므로 고객사 실명은 시드에 남기지 않는다.
#: .gitignore가 원본 xlsx(sources/)를 제외하는 것과 같은 방침을, 원본에서 파생된
#: 시드에도 적용한다. 룰 로직·코드값·임계값은 그대로 두므로 목업 검증 목적은 유지된다.
#:
#: 실명↔별칭 대응표 자체가 실명을 담으므로 이 파일에 두지 않고 sources/mask_map.json
#: (gitignore 대상)에서 읽는다. 형식:
#:   {"names": [["실명", "수요가 A"], ...], "codes": {"110141": "수요가 A", ...}}
#: 맵이 없으면 마스킹할 실명을 모르므로, 안전한 쪽으로 동작한다 —
#: 고객 필드(FNL_CUS_CD 등)의 값 라벨을 통째로 비우고, 메시지는 원문 그대로 두되
#: --check 어서션이 "마스킹 맵 없음"을 실패로 잡아 공개 빌드를 막는다.
MASK_MAP_PATH = ROOT / "sources" / "mask_map.json"
CUSTOMER_FIELDS = ("FNL_CUS_CD", "CUS_CD", "ACT_CUS_CD")


def load_mask_map():
    if not MASK_MAP_PATH.exists():
        return None
    try:
        m = json.loads(MASK_MAP_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError) as e:
        raise PipelineError(f"마스킹 맵을 읽을 수 없습니다 ({MASK_MAP_PATH}): {e}")
    names = [(str(a), str(b)) for a, b in m.get("names", [])]
    # 긴 이름부터 치환해야 부분 문자열이 먼저 먹지 않는다
    names.sort(key=lambda p: -len(p[0]))
    return {"names": names, "codes": {str(k): str(v) for k, v in m.get("codes", {}).items()}}


def mask_text(s, mm):
    """에러 메시지·근거 문자열에서 고객사 실명을 별칭으로 치환한다."""
    if not isinstance(s, str) or not mm:
        return s
    for name, alias in mm["names"]:
        s = s.replace(name, alias)
    return s


def apply_masking(seed, mm):
    """시드 전체(룰 메시지·값 라벨·이슈·계보 원문)에서 고객사 실명을 제거한다."""
    def mask(s):
        return mask_text(s, mm)
    n_msg = n_label = 0
    for r in seed["rules"]:
        err = r.get("error", {})
        for key in ("message", "raw_message"):
            if isinstance(err.get(key), str):
                masked = mask(err[key])
                if masked != err[key]:
                    err[key] = masked
                    if key == "message":
                        n_msg += 1
        for lin in r.get("lineage", []):
            for key in ("raw_message", "diff"):
                if key in lin:
                    lin[key] = mask(lin[key])
    labels = []
    for v in seed.get("value_labels", []):
        if v.get("field") in CUSTOMER_FIELDS:
            continue  # 고객 라벨은 아래에서 별칭으로만 재생성한다 (실명 노출 금지)
        labels.append(v)
    # 룰이 실제로 참조하는 고객 코드에만 별칭 라벨을 붙인다 — 화면에서 "110141"이
    # 무엇인지 알 수 있게 하되, 실명은 저장소 밖(sources/mask_map.json)에만 둔다.
    used = set()
    for r in seed["rules"]:
        for c in r.get("when", []) + r.get("context", []):
            if c.get("field") in CUSTOMER_FIELDS:
                used.update(str(x) for x in c.get("values", []))
    for code, alias in sorted((mm or {}).get("codes", {}).items()):
        if code not in used:
            continue
        labels.append({"field": "FNL_CUS_CD", "code": code, "label": alias,
                       "certainty": "추정",
                       "evidence": "공개 배포용 별칭 — 실명은 저장소에 포함하지 않는다"})
        n_label += 1
    seed["value_labels"] = sorted(labels, key=lambda x: (x["field"], str(x["code"])))
    for i in seed.get("issues", []):
        blob = json.dumps(i, ensure_ascii=False)
        masked = mask(blob)
        if masked != blob:
            i.clear()
            i.update(json.loads(masked))
    seed["stats"]["masking"] = {
        "applied": True,
        "reason": "공개 저장소 배포 — 고객사 실명 비노출 (.gitignore sources/ 방침과 동일)",
        "messages_masked": n_msg,
        "customer_labels_aliased": n_label,
        "alias_count": len(set(a for _, a in (mm or {}).get("names", []))),
    }
    return seed


def build_seed(order_path, cross_path, mask=True):
    stats = {"display_hints": [], "in_demoted": 0, "not_in_demoted": 0}
    parsed_order = parse_workbook(order_path)
    parsed_cross = parse_workbook(cross_path)
    labels = derive_labels({"order": parsed_order, "cross": parsed_cross})
    order_clean = clean_rules(parsed_order, "order", labels, stats)
    cross_clean = clean_rules(parsed_cross, "cross", labels, stats)

    # ---- Task 1 Step 4: --check 어서션 (a)~(f) ----
    checks = []
    def check(name, cond, detail=""):
        checks.append((name, cond, detail))
        if not cond:
            raise PipelineError(f"어서션 실패: {name} {detail}")
    check("(a) 실질 룰 수 order=136", len(order_clean) == 136, f"실제 {len(order_clean)}")
    check("(a) 실질 룰 수 cross=98", len(cross_clean) == 98, f"실제 {len(cross_clean)}")
    # (b)(c)(d)(e)는 clean_condition/파서에서 위반 시 즉시 PipelineError — 통과 자체가 증거
    check("(b) IS_NULL/NOT_NULL에 값 존재 0건", True, "clean_condition에서 강제")
    check("(c) BETWEEN low>high 0건", True, "clean_condition에서 강제")
    check("(d) 숫자 op 비숫자 잔존 0건", True, "_num에서 강제")
    check("(e) 조건 0개 룰 0건", True, "parse_workbook에서 강제")
    dept_order = [r["owner_dept"] for r in order_clean]
    check("(f) order 태그 [품질설계] 133건", dept_order.count("품질설계") == 133,
          f"실제 {dept_order.count('품질설계')}")
    check("(f) order 태그 [OMS] 3건", dept_order.count("OMS") == 3, f"실제 {dept_order.count('OMS')}")
    check("(f) cross 태그 0건", all(r["owner_dept"] is None for r in cross_clean))
    trunc = sorted((r["source"], r["row"]) for r in order_clean + cross_clean
                   if "truncated_message" in r["flags"])
    check("끝 콤마(truncated) = order#135,#136", trunc == [("order", 135), ("order", 136)], str(trunc))
    # 숫자 승격 검증: 숫자 연산자 조건의 values/low/high가 전부 number 타입인지
    for r in order_clean + cross_clean:
        for c in r["when"] + r["context"]:
            if c["op"] in ("GT", "GE", "LT", "LE"):
                assert all(isinstance(v, (int, float)) and not isinstance(v, bool)
                           for v in c["values"]), (r["source"], r["row"], c)
            if c["op"] == "BETWEEN":
                assert isinstance(c["low"], (int, float)) and isinstance(c["high"], (int, float))
    check("(d+) 숫자 조건 float 승격 전수 확인", True)

    # ---- Task 2: 병합 ----
    rules, cross85, detail = build_ruleset(order_clean, cross_clean, labels, stats)
    issues = build_issues(rules, cross85, detail, labels)
    fields = build_fields(rules, labels, parsed_order, parsed_cross)

    # ---- Task 2 Step 5: 병합 어서션 ----
    n_total = len(rules)
    n_common = sum(1 for r in rules if r["source"] == "공통")
    n_order_only = sum(1 for r in rules if r["source"] == "order")
    n_cross_only = sum(1 for r in rules if r["source"] == "cross")
    needs_review_ids = sorted(r["id"] for r in rules if r["merge"]["status"] == "needs_review")
    conflict_ids = [r["id"] for r in rules if r["merge"]["status"] == "conflict_resolved"]
    check("병합: 공통(조건 시그니처 완전 동일) 77", n_common == 77, f"실제 {n_common}")
    check("병합: order 채택 59 (고유 58 + M3 부분집합 흡수 1 = R-081)",
          n_order_only == 59, f"실제 {n_order_only}")
    check("병합: cross 고유 탑재 2 (R-C50, R-C98)", n_cross_only == 2, f"실제 {n_cross_only}")
    check("병합 합산식: 77+59+1+1 = 총 룰 수", 77 + 59 + 1 + 1 == n_total, f"총 {n_total}")
    # source='공통' 배지는 "cross 계보 행과 조건이 완전 동일"을 주장한다 — 전건 데이터로 재검증.
    cross_by_row = {c["row"]: c for c in cross_clean}
    common_mismatch = []
    for r in rules:
        if r["source"] != "공통":
            continue
        my = rule_sig(r)
        for l in r["lineage"]:
            if l["source"] != "cross":
                continue
            if rule_sig(cross_by_row[l["row"]]) != my:
                common_mismatch.append((r["id"], f"cross#{l['row']}"))
    check("source='공통' 전건이 cross 계보 행과 조건 완전 동일",
          not common_mismatch, str(common_mismatch))
    check("확인필요 정확히 6건",
          needs_review_ids == ["R-073", "R-075", "R-084", "R-088", "R-095", "R-C50"],
          str(needs_review_ids))
    check("conflict_resolved ≥ 1 (R-010)", conflict_ids == ["R-010"], str(conflict_ids))
    # cross 98행 계보 회계 — 산술 선언이 아니라 실제 산출물에서 역산해 전건 보존을 확인한다.
    cross_rows_lineage = {l["row"] for r in rules for l in r["lineage"] if l["source"] == "cross"}
    cross_rows_issues = {s["row"] for i in issues
                         for s in (i.get("left"), i.get("right"))
                         if s and s.get("source") == "cross" and s.get("row") is not None}
    covered = cross_rows_lineage | cross_rows_issues
    check("cross 98행 전건 계보 보존 (lineage ∪ issues)",
          covered == set(range(1, 99)),
          f"미보존 {sorted(set(range(1, 99)) - covered)}")
    check("cross#85(A451 >8000)는 룰 미탑재 + lineage(R-098)·이슈(I-C3) 양쪽 보존",
          all("R-C85" != r["id"] for r in rules)
          and cross85["err_code"] == "A451"
          and 85 in cross_rows_lineage and 85 in cross_rows_issues)
    check("이슈 수 = 상충3 + 확인필요6 + 데이터결함5 = 14", len(issues) == 14, f"실제 {len(issues)}")
    n_review_alt = sum(1 for r in rules if r["merge"].get("review_alternatives"))
    check("cross 대안값 병기 5건(M5)", n_review_alt == 5, f"실제 {n_review_alt}")
    field_keys = {f["key"] for f in fields}
    for r in rules:
        for c in r["when"] + r["context"]:
            assert c["field"] in field_keys, (r["id"], c["field"])
        assert r["target"]["item"] in field_keys, r["id"]
    check("모든 조건 필드가 사전에 등재", True)
    for vl in VALUE_LABELS:
        assert vl["field"] in field_keys, vl
        assert vl["certainty"] == "추정", vl
    check("값 라벨 전건 '추정' + 필드 등재 확인", True)

    unused_order_cols = [c for c in parsed_order["fields"]
                         if c not in ("주문항목코드", "주문항목값")
                         and not any(cc["field"] == c for r in parsed_order["rules"] for cc in r["conds"])]
    unused_cross_cols = [c for c in parsed_cross["fields"]
                         if c not in ("주문항목코드", "주문항목값")
                         and not any(cc["field"] == c for r in parsed_cross["rules"] for cc in r["conds"])]

    stats.update({
        "input": {
            # 파일명만 기록한다 — 절대경로를 시드에 박으면 머신마다 산출물이 달라져
            # content_hash·주입 페이로드가 재현되지 않는다(입력 경로는 실행 옵션이지 데이터가 아님).
            "order": {"file": Path(order_path).name, "rules": len(order_clean), **parsed_order["meta"]},
            "cross": {"file": Path(cross_path).name, "rules": len(cross_clean), **parsed_cross["meta"]},
        },
        "merge": {
            "sig_pairs": stats.pop("sig_pairs"),
            "common": n_common, "order_only": n_order_only, "cross_mounted": n_cross_only,
            "drift_expansion_M4": stats.pop("drift_expansion"),
            "drift_reduction_M5": stats.pop("drift_reduction"),
            "conflict_resolved_M2": len(conflict_ids),
            "absorbed_M3": 1, "complement_M8": 1, "cross_discarded_M6": 1,
            "formula": f"공통(조건 동일) {n_common} + order 채택 {n_order_only}"
                       f"(고유 58 + M3 부분집합 흡수 1) + cross 고유 탑재 {n_cross_only}"
                       f"(A321 비활성 1 + 조합폭10 IS_NULL 상보 1) = 통합 {n_total}룰 "
                       f"(cross#85 A451 구룰은 룰 미탑재 — R-098 lineage + 이슈 I-C3에 보존)",
            "cross_accounting": "cross 98 = 시그니처 동일 78(그중 77은 '공통' 룰의 계보, "
                                "1은 (order#77,cross#69) 재짝짓기로 cross#69→R-081 흡수) "
                                "+ 드리프트 계보 17(확장12·축소5 — cross#65→R-077 포함) "
                                "+ 탑재 2(R-C50·R-C98) + 폐기·계보보존 1(cross#85 → R-098 lineage·이슈 I-C3) "
                                "= 77+1+17+2+1",
            "sig_pair_repaired": {"order_row": 77, "cross_row": 69,
                                  "note": "조건 시그니처 완전 동일 쌍이지만 코드·메시지 상충(A72 vs A722)이라 "
                                          "재짝짓기 — '공통' 배지 대상에서 제외(이슈 I-C2)"},
        },
        "total_rules": n_total,
        "needs_review": len(needs_review_ids),
        "needs_review_ids": needs_review_ids,
        "conflicts": sum(1 for i in issues if i["type"] == "conflict"),
        "issues": len(issues),
        "issues_by_type": {t: sum(1 for i in issues if i["type"] == t)
                           for t in ("conflict", "needs_review", "data_defect")},
        "fields": len(fields),
        "value_labels": len(VALUE_LABELS),
        "categories": {cat: sum(1 for r in rules if r["error"]["category"] == cat)
                       for cat in CATEGORY_ORDER},
        "severity_candidate_warn": sum(1 for r in rules if r["error"].get("severity_candidate") == "warn"),
        "unused_columns": {"order": unused_order_cols, "cross": unused_cross_cols},
        "flags": {
            "co_fire_pair": sorted(r["id"] for r in rules if "co_fire_pair" in r["flags"]),
            "truncated_message": sorted(r["id"] for r in rules if "truncated_message" in r["flags"]),
        },
    })

    seed = {
        "schema_version": "1.0",
        "ruleset": {
            "id": "order-check-merged",
            "title": "주문 정합성 체크 통합 룰셋",
            "sources": [
                {"key": "order", "file": "주문에러체크.xlsx", **parsed_order["meta"]},
                {"key": "cross", "file": "항목간주문에러체크.xlsx", **parsed_cross["meta"]},
            ],
            "semantics": {
                "evaluation": "전 룰 평가·발화 전수 수집 (first-match 금지, expert2 §4.1)",
                "and": "when·context는 평가상 전부 AND — 구분은 문장화 표현용",
                "between": "BETWEEN1=[low,high] 폐구간 / BETWEEN2=[low,high) — lowInclusive/highInclusive로 정규화",
                "null": "미입력·빈문자열·null 동일 취급, IS_NULL만 참. 비교연산은 NULL 피연산자면 미발화. 0 ≠ NULL",
                "source_values": {
                    "공통": "order·cross 조건 시그니처 완전 동일(계보 병합) — 조건이 다르면 이 배지를 주지 않는다",
                    "order": "order 고유 · 드리프트 채택 · 부분집합 흡수(R-081) — cross 계보가 있어도 조건은 order 값",
                    "cross": "cross 고유 탑재(R-C50·R-C98)"},
            },
        },
        "fields": fields,
        "value_labels": VALUE_LABELS,
        "rules": rules,
        "issues": issues,
        "stats": stats,
    }
    if mask:
        mm = load_mask_map()
        if mm is None:
            raise PipelineError(
                f"마스킹 맵이 없습니다: {MASK_MAP_PATH}\n"
                "  공개 저장소 배포용 시드는 고객사 실명을 담을 수 없습니다.\n"
                '  형식: {"names": [["실명","수요가 A"]], "codes": {"110141":"수요가 A"}}\n'
                "  사내 배포용으로 실명을 살리려면 --no-mask 를 쓰세요(공개 저장소에는 금지).")
        apply_masking(seed, mm)
        blob = json.dumps(seed, ensure_ascii=False)
        leaked = [n for n, _ in mm["names"] if n in blob]
        check("마스킹: 고객사 실명 잔존 0건", not leaked, ", ".join(leaked) or "없음")
    else:
        seed["stats"]["masking"] = {"applied": False, "reason": "--no-mask (사내 배포용)"}

    # 시드 '내용' 해시 — 모듈 loadState()의 재시드 게이트. schema_version/LS_KEY는 리터럴이라
    # 룰이 바뀌어도 값이 그대로여서, 이전 빌드를 연 브라우저가 구 룰셋에 영구 고정되던 문제를 막는다.
    seed["content_hash"] = content_hash(seed)
    check("content_hash 생성(12자리 sha1)",
          len(seed["content_hash"]) == 12 and seed["content_hash"] == content_hash(seed),
          seed["content_hash"])
    return seed, checks


def content_hash(seed):
    payload = json.dumps(
        {k: seed[k] for k in ("rules", "issues", "fields", "value_labels")},
        ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()[:12]


# ---------------------------------------------------------------------------
# --inject (멱등)
# ---------------------------------------------------------------------------
def inject(seed):
    if not MODULE_HTML.exists():
        print(f"경고: {MODULE_HTML} 없음 — 주입 skip (모듈 생성 후 --inject 재실행)", file=sys.stderr)
        return False
    html = MODULE_HTML.read_text(encoding="utf-8")
    if MARKER_START not in html or MARKER_END not in html:
        print(f"경고: {MODULE_HTML}에 시드 마커({MARKER_START} ~ {MARKER_END}) 없음 — 주입 skip",
              file=sys.stderr)
        return False
    payload = json.dumps(seed, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    block = f"{MARKER_START}\nconst OC_SEED = {payload};\n{MARKER_END}"
    pattern = re.compile(re.escape(MARKER_START) + r".*?" + re.escape(MARKER_END), re.DOTALL)
    if len(pattern.findall(html)) != 1:
        raise PipelineError("시드 마커 구간이 1개가 아님")
    new_html = pattern.sub(lambda _: block, html)
    # 멱등성 assert: 다시 치환해도 동일
    again = pattern.sub(lambda _: block, new_html)
    assert again == new_html, "주입 멱등성 위반"
    if new_html != html:
        MODULE_HTML.write_text(new_html, encoding="utf-8")
        print(f"주입 완료: {MODULE_HTML} ({len(payload):,} bytes seed)")
    else:
        print(f"주입 생략: {MODULE_HTML} 이미 최신 (멱등 확인)")
    return True


def main():
    ap = argparse.ArgumentParser(description="주문정합성체크 정제 파이프라인")
    ap.add_argument("--order", default=DEFAULT_ORDER,
                    help=f"주문에러체크.xlsx 경로 (기본: OC_XLSX_DIR 또는 sources/ — 현재 {DEFAULT_ORDER})")
    ap.add_argument("--cross", default=DEFAULT_CROSS,
                    help=f"항목간주문에러체크.xlsx 경로 (기본: OC_XLSX_DIR 또는 sources/ — 현재 {DEFAULT_CROSS})")
    ap.add_argument("--check", action="store_true", help="어서션 결과 출력 (실패 시 exit 1)")
    ap.add_argument("--emit", action="store_true", help=f"{SEED_OUT} 생성")
    ap.add_argument("--inject", action="store_true", help="모듈 마커에 시드 주입 (멱등)")
    ap.add_argument("--no-mask", action="store_true",
                    help="고객사 실명 마스킹 해제 (사내 배포용 — 공개 저장소에는 쓰지 말 것)")
    args = ap.parse_args()

    try:
        seed, checks = build_seed(args.order, args.cross, mask=not args.no_mask)
    except PipelineError as e:
        print(f"FAIL: {e}", file=sys.stderr)
        sys.exit(1)

    if args.check:
        for name, ok, detail in checks:
            print(f"  PASS {name}" + (f" ({detail})" if detail else ""))
        print(f"어서션 {len(checks)}건 전부 통과")
    if args.emit:
        SEED_OUT.write_text(json.dumps(seed, ensure_ascii=False, indent=1), encoding="utf-8")
        print(f"시드 생성: {SEED_OUT} ({SEED_OUT.stat().st_size:,} bytes)")
    if args.inject:
        inject(seed)
    if not (args.check or args.emit or args.inject):
        print(json.dumps(seed["stats"], ensure_ascii=False, indent=2))
    s = seed["stats"]
    print(f"요약: 총 {s['total_rules']}룰 ({s['merge']['formula']}) · "
          f"확인필요 {s['needs_review']} · 이슈 {s['issues']} · "
          f"필드 {s['fields']} · 값라벨 {s['value_labels']}")


if __name__ == "__main__":
    main()
